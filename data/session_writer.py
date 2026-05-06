"""
Session file management — names, folders, and the post-session summary CSV.
"""

from __future__ import annotations

import csv
import math
import re
from pathlib import Path
from typing import Optional

from core.event_log import Event, EventLog, EventType
from core.session import Session


def resolve_csv_path(session: Session) -> Path:
    """
    Return the full path for the session CSV, appending _2, _3 etc.
    if the filename already exists to prevent overwriting.
    """
    folder = Path(session.csv_folder())
    folder.mkdir(parents=True, exist_ok=True)
    base = session.csv_filename()
    stem = Path(base).stem
    suffix = Path(base).suffix

    candidate = folder / base
    counter = 2
    while candidate.exists():
        candidate = folder / f"{stem}_{counter}{suffix}"
        counter += 1
    return candidate


def build_metadata(
    session: Session,
    stimulus_set,
    counterbalance_notes: str = "",
) -> dict[str, str]:
    """Build the metadata dictionary written as # comments at the top of the CSV."""
    p = session.current_stim_params
    included = " | ".join(stimulus_set.included_ids[:30])
    if len(stimulus_set.included_ids) > 30:
        included += f" … (+{len(stimulus_set.included_ids) - 30} more)"
    excluded = " | ".join(stimulus_set.excluded_ids) or "none"

    return {
        "SessionID":               session.session_id,
        "SoftwareVersion":         _get_version(),
        "PatientID":               session.patient_id,
        "SessionDate":             session.session_date,
        "SessionStartTime":        session.session_start_time,
        "TestType":                session.task_code,
        "Electrode":               p.electrode,
        "Contact":                 p.contact,
        "StimulationIntensity_mA": str(p.intensity_ma),
        "StimulationDuration_s":   str(p.duration_s),
        "ProgressionMode":         session.progression_mode,
        "StimSignalKey":           session.stim_key,
        "StimuliOrder":            session.stimuli_order,
        "ScreenWidth_px":          str(session.screen_width),
        "ScreenHeight_px":         str(session.screen_height),
        "StimuliIncluded":         included,
        "StimuliExcluded":         excluded,
        "CounterBalance":          counterbalance_notes,
    }


def write_summary(
    session:     Session,
    event_log:   EventLog,
    n_trials:    int,
    n_correct:   int,
    n_skipped:   int,
) -> Path:
    """
    Write the post-session summary CSV and return its path.
    """
    folder = Path(session.csv_folder())
    folder.mkdir(parents=True, exist_ok=True)
    summary_path = folder / session.summary_filename()

    response_events = [
        ev for ev in event_log.events if ev.event == EventType.RESPONSE
    ]
    trs = [ev.tr_s for ev in response_events if ev.tr_s is not None]
    mean_tr = sum(trs) / len(trs) if trs else None
    sd_tr   = (
        math.sqrt(sum((t - mean_tr) ** 2 for t in trs) / len(trs))
        if len(trs) > 1 else None
    )

    # Per-epoch breakdown
    n_stim_events = sum(
        1 for ev in event_log.events if ev.event == EventType.STIM_START
    )

    def _epoch_stats(label: str):
        evs = [ev for ev in response_events if ev.stim_epoch == label]
        n   = len(evs)
        ok  = sum(1 for ev in evs if ev.correct == "Yes")
        trs_ep = [ev.tr_s for ev in evs if ev.tr_s is not None]
        mtr = round(sum(trs_ep) / len(trs_ep), 4) if trs_ep else None
        return n, ok, mtr

    n_pre,  ok_pre,  mtr_pre  = _epoch_stats("pré-stim")
    n_per,  ok_per,  mtr_per  = _epoch_stats("per-stim")
    n_lim,  ok_lim,  mtr_lim  = _epoch_stats("limite-stim")
    n_post, ok_post, mtr_post = _epoch_stats("post-stim")

    # When no stimulation occurred, per/limite/post stats are meaningless — write None
    if n_stim_events == 0:
        n_per = ok_per = mtr_per = None
        n_lim = ok_lim = mtr_lim = None
        n_post = ok_post = mtr_post = None

    def _fmt(v):
        return v if v is not None else ""

    with summary_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow([
            "task", "n_trials", "n_correct",
            "mean_TR_s", "sd_TR_s", "n_timeout", "n_skipped",
            "n_stim_events",
            "n_trials_pre",    "n_correct_pre",    "mean_TR_pre",
            "n_trials_per",    "n_correct_per",    "mean_TR_per",
            "n_trials_limite", "n_correct_limite", "mean_TR_limite",
            "n_trials_post",   "n_correct_post",   "mean_TR_post",
        ])
        writer.writerow([
            session.task_code,
            n_trials,
            n_correct,
            round(mean_tr, 4) if mean_tr is not None else "",
            round(sd_tr,   4) if sd_tr   is not None else "",
            0,
            n_skipped,
            n_stim_events,
            _fmt(n_pre),  _fmt(ok_pre),  _fmt(mtr_pre),
            _fmt(n_per),  _fmt(ok_per),  _fmt(mtr_per),
            _fmt(n_lim),  _fmt(ok_lim),  _fmt(mtr_lim),
            _fmt(n_post), _fmt(ok_post), _fmt(mtr_post),
        ])

    return summary_path


def _get_version() -> str:
    try:
        import config
        return config.SOFTWARE_VERSION
    except Exception:
        return "unknown"


# ─── Human-readable stimulus labels ──────────────────────────────────────────

def _format_stimulus_label(stim_id: str, task_code: str) -> str:
    """Convert a raw planche_id to a human-readable label."""
    if not stim_id:
        return stim_id
    task = task_code.upper()

    if task in ("ASM_MOTS", "ASM_SEEG"):
        # "AS_SEEGsansDenoV2_scie_branche" → "scie / branche" (last 2 underscore parts)
        parts = stim_id.split("_")
        if len(parts) >= 2:
            return f"{parts[-2]} / {parts[-1]}"
        return stim_id

    if task in ("FFP_V1", "FFP_V2"):
        # "FFP_V1_ZinedineZidane" → "Zinedine Zidane"
        m = re.match(r'^FFP_V\d_(.+)$', stim_id, re.IGNORECASE)
        if m:
            return re.sub(r'(?<=[a-z])(?=[A-Z])', ' ', m.group(1))
        return stim_id

    if task in ("MUF_V1", "MUF_V2"):
        # "MFU_V1_targetLEFT_06" → "Visage inconnu 06 (gauche)"
        m = re.match(r'.*_target(LEFT|RIGHT|CENTRE|CENTER)_(\d+)$', stim_id, re.IGNORECASE)
        if m:
            side_map = {"LEFT": "gauche", "RIGHT": "droite", "CENTRE": "centre", "CENTER": "centre"}
            side = side_map.get(m.group(1).upper(), m.group(1).lower())
            return f"Visage inconnu {m.group(2)} ({side})"
        return stim_id

    if task == "DI_SEEG":
        # "DI_SEEG_sandales_125" → "sandales"
        m = re.match(r'^DI_SEEG_(.+?)(?:_\d+)?$', stim_id, re.IGNORECASE)
        if m:
            return m.group(1)
        return stim_id

    if task == "FNP":
        # "FNP_Patrick_Bruel" → "Patrick Bruel"
        m = re.match(r'^FNP_(.+)$', stim_id, re.IGNORECASE)
        if m:
            return m.group(1).replace('_', ' ')
        return stim_id

    return stim_id


def _expected_response(stim_id: str, task_code: str) -> str:
    """Extract expected/target response from stimulus ID."""
    if not stim_id:
        return ""
    task = task_code.upper()
    if task in ("MUF_V1", "MUF_V2"):
        m = re.match(r'.*_target(LEFT|RIGHT|CENTRE|CENTER)_', stim_id, re.IGNORECASE)
        if m:
            side_map = {"LEFT": "gauche", "RIGHT": "droite",
                        "CENTRE": "centre", "CENTER": "centre"}
            return side_map.get(m.group(1).upper(), m.group(1).lower())
    return ""


# ─── Excel report ─────────────────────────────────────────────────────────────

def write_excel_report(
    session:   Session,
    event_log: EventLog,
    n_trials:  int,
    n_correct: int,
    n_skipped: int,
) -> Path:
    """Write a multi-sheet Excel report and return its path."""
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers as xl_numbers
    )
    from openpyxl.utils import get_column_letter

    folder = Path(session.csv_folder())
    folder.mkdir(parents=True, exist_ok=True)
    date  = __import__("datetime").datetime.now().strftime("%Y-%m-%d")
    hhmm  = __import__("datetime").datetime.now().strftime("%H-%M")
    p     = session.current_stim_params
    mA    = str(p.intensity_ma).rstrip("0").rstrip(".")
    s_dur = str(p.duration_s).rstrip("0").rstrip(".")
    fname = (
        f"Patient_{session.patient_id}_{date}_{hhmm}_{session.task_code}"
        f"_Contact{p.electrode}-{p.contact}_{mA}mA_{s_dur}s_rapport.xlsx"
    )
    xl_path = folder / fname

    wb = openpyxl.Workbook()

    # ── Style helpers ────────────────────────────────────────────────────────
    def _fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color.lstrip("#"))

    def _font(bold=False, color="000000", size=11) -> Font:
        return Font(bold=bold, color=color, size=size)

    def _center() -> Alignment:
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _left() -> Alignment:
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin")
    def _border() -> Border:
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    events     = event_log.events
    task_code  = session.task_code
    resp_evs   = [ev for ev in events if ev.event == EventType.RESPONSE]
    trs_all    = [ev.tr_s for ev in resp_evs if ev.tr_s is not None]
    mean_tr    = sum(trs_all) / len(trs_all) if trs_all else None
    n_stim     = sum(1 for ev in events if ev.event == EventType.STIM_START)

    def _ep(label):
        evs = [ev for ev in resp_evs if ev.stim_epoch == label]
        n   = len(evs)
        ok  = sum(1 for ev in evs if ev.correct == "Yes")
        trs = [ev.tr_s for ev in evs if ev.tr_s is not None]
        mtr = round(sum(trs)/len(trs), 4) if trs else None
        return n, ok, mtr

    n_pre, ok_pre, mtr_pre = _ep("pré-stim")
    n_per, ok_per, mtr_per = _ep("per-stim")
    n_lim, ok_lim, mtr_lim = _ep("limite-stim")
    n_post,ok_post,mtr_post= _ep("post-stim")

    # ── Sheet 1: Résumé (full clinical summary) ──────────────────────────────
    import statistics as _statistics

    ws1 = wb.active
    ws1.title = "Résumé"
    ws1.sheet_view.showGridLines = False

    hdr_fill = _fill("1a1a2e")   # kept for later sheets

    # ── Compute additional stats ──────────────────────────────────────────────
    ts_first  = events[0].time_s  if events else 0.0
    ts_last   = events[-1].time_s if events else 0.0
    dur_secs  = max(0.0, ts_last - ts_first)
    dur_str   = f"{int(dur_secs // 60)}m {int(dur_secs % 60)}s"

    tr_median = round(_statistics.median(trs_all), 3) if trs_all else None
    tr_min    = round(min(trs_all), 3)                if trs_all else None
    tr_max    = round(max(trs_all), 3)                if trs_all else None

    pct_correct = round(n_correct / n_trials * 100, 1) if n_trials else 0.0
    acc_color   = ("16a34a" if pct_correct >= 70
                   else ("d97706" if pct_correct >= 40 else "dc2626"))
    n_incorrect = sum(1 for ev in resp_evs if ev.correct == "No")

    all_trial_essais = {ev.essai for ev in events
                        if ev.event == EventType.TRIAL_START and ev.essai is not None}
    responded_essais = {ev.essai for ev in events
                        if ev.event == EventType.RESPONSE and ev.essai is not None}
    skipped_essais   = {ev.essai for ev in events
                        if ev.event == EventType.STIMULUS_SKIP and ev.essai is not None}
    excl_essais      = {ev.essai for ev in events
                        if ev.event == EventType.STIMULUS_EXCLUDE and ev.essai is not None}
    n_exclu          = len(excl_essais)
    n_timeout        = len(all_trial_essais - responded_essais - skipped_essais - excl_essais)

    stim_detail_windows: list[dict] = []
    _open_sd: Optional[dict] = None
    for ev in events:
        if ev.event == EventType.STIM_START:
            _open_sd = {"start_s": ev.time_s, "start_iso": ev.time_iso,
                        "notes": ev.notes or ""}
        elif ev.event == EventType.STIM_END and _open_sd is not None:
            _open_sd.update(end_s=ev.time_s, end_iso=ev.time_iso)
            stim_detail_windows.append(_open_sd)
            _open_sd = None
    if _open_sd is not None:
        last_ev = events[-1] if events else None
        _open_sd.update(end_s=last_ev.time_s if last_ev else 0.0,
                        end_iso=last_ev.time_iso if last_ev else "")
        stim_detail_windows.append(_open_sd)

    clinician_notes = ""
    for ev in events:
        if ev.event == EventType.SESSION_NOTES:
            clinician_notes = ev.notes or ""
            break

    # ── Sheet 1 row-counter and layout helpers ────────────────────────────────
    _s1_r = [0]

    def _nr() -> int:
        _s1_r[0] += 1
        return _s1_r[0]

    C1_NAVY  = "1a2744"
    C1_LABEL = "6b7280"
    C1_VALUE = "111318"

    def _s1_section_hdr(text: str) -> None:
        r = _nr()
        ws1.row_dimensions[r].height = 20
        c = ws1.cell(r, 1, text)
        c.font      = Font(bold=True, color="FFFFFF", size=11)
        c.fill      = PatternFill("solid", fgColor=C1_NAVY)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

    def _s1_kv(label: str, value, value_bold: bool = False,
               value_color: str = "") -> None:
        r = _nr()
        ws1.row_dimensions[r].height = 16
        la = ws1.cell(r, 1, label)
        la.font      = Font(size=11, color=C1_LABEL)
        la.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        vl = ws1.cell(r, 2, value)
        vl.font      = Font(size=11, color=value_color or C1_VALUE, bold=value_bold)
        vl.alignment = Alignment(horizontal="left", vertical="center")

    def _s1_spacer() -> None:
        r = _nr()
        ws1.row_dimensions[r].height = 8
        ws1.cell(r, 1, " ")   # empty placeholder so max_row advances

    # ── Section 1: Session info ───────────────────────────────────────────────
    _s1_section_hdr("RÉSUMÉ DE SESSION")
    _s1_kv("Patient",          session.patient_id)
    _s1_kv("Tâche",            session.task_display_name)
    _s1_kv("Date",             session.session_date)
    _s1_kv("Heure début",      session.session_start_time)
    _s1_kv("Durée session",    dur_str)
    _s1_kv("Électrode",        p.electrode)
    _s1_kv("Contact",          p.contact)
    _s1_kv("Intensité",        f"{p.intensity_ma} mA")
    _s1_kv("Durée stim",       f"{p.duration_s} s")
    _s1_kv("Mode progression", session.progression_mode)
    _s1_kv("Version",          _get_version())

    # ── Section 2: Global results ─────────────────────────────────────────────
    _s1_spacer()
    _s1_section_hdr("RÉSULTATS GLOBAUX")
    _s1_kv("Essais total",  n_trials,   value_bold=True)
    _s1_kv(
        "Corrects",
        f"{n_correct}  ({pct_correct}%)",
        value_bold=True, value_color=acc_color,
    )
    n_inc_pct = round(n_incorrect / n_trials * 100, 1) if n_trials else 0
    _s1_kv("Incorrects",    f"{n_incorrect}  ({n_inc_pct}%)", value_bold=True)
    _s1_kv("TR moyen",  f"{round(mean_tr, 2)} s"  if mean_tr   is not None else "—")
    _s1_kv("TR médian", f"{tr_median} s"           if tr_median is not None else "—")
    _s1_kv("TR min",    f"{tr_min} s"              if tr_min    is not None else "—")
    _s1_kv("TR max",    f"{tr_max} s"              if tr_max    is not None else "—")
    _s1_kv("Passés",    n_skipped,  value_bold=True)
    _s1_kv("Exclus",    n_exclu,    value_bold=True)
    _s1_kv("Timeouts",  n_timeout,  value_bold=True)

    # ── Section 3: Stimulations ───────────────────────────────────────────────
    if n_stim > 0:
        _s1_spacer()
        _s1_section_hdr("STIMULATIONS")
        _s1_kv("Événements STIM", n_stim, value_bold=True)
        _s1_spacer()

        # Epoch sub-table header
        r_ep = _nr()
        ws1.row_dimensions[r_ep].height = 16
        for ci, txt in enumerate(
            ["Époque", "N essais", "Corrects (%)", "TR moyen (s)"], start=1
        ):
            c = ws1.cell(r_ep, ci, txt)
            c.font      = Font(bold=True, size=10, color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor="334155")
            c.alignment = Alignment(horizontal="center", vertical="center")

        # Epoch data rows
        epoch_rows = [
            ("Pré-stim",    n_pre,  ok_pre,  mtr_pre,  "FFFFFF", False),
            ("Per-stim",    n_per,  ok_per,  mtr_per,  "fee2e2", True),
            ("Limite stim", n_lim,  ok_lim,  mtr_lim,  "fff3e0", True),
            ("Post-stim",   n_post, ok_post, mtr_post, "e0f2fe", False),
        ]
        for ep_name, ep_n, ep_ok, ep_mtr, ep_hex, ep_bold in epoch_rows:
            r_ep = _nr()
            ws1.row_dimensions[r_ep].height = 16
            ep_fill = PatternFill("solid", fgColor=ep_hex)
            ep_pct  = round(ep_ok / ep_n * 100, 1) if ep_n else 0
            ep_vals = [
                ep_name,
                ep_n if ep_n else 0,
                f"{ep_ok}  ({ep_pct}%)" if ep_n else "—",
                f"{ep_mtr} s" if ep_mtr is not None else "—",
            ]
            for ci, val in enumerate(ep_vals, start=1):
                c = ws1.cell(r_ep, ci, val)
                c.font      = Font(size=10, bold=ep_bold, color=C1_VALUE)
                c.fill      = ep_fill
                c.alignment = Alignment(horizontal="center", vertical="center")

        # Stim detail blocks
        if stim_detail_windows:
            _s1_spacer()
            _s1_section_hdr("STIMULATIONS DÉTAIL")
            for wi, wd in enumerate(stim_detail_windows):
                elec_m = re.search(r'Electrode=(\S+)', wd["notes"])
                cont_m = re.search(r'Contact=(\S+)',   wd["notes"])
                inty_m = re.search(r'Intensity=(\S+)', wd["notes"])
                dura_m = re.search(r'Duration=(\S+)',  wd["notes"])
                elec_v = elec_m.group(1) if elec_m else p.electrode
                cont_v = cont_m.group(1) if cont_m else p.contact
                inty_v = inty_m.group(1) if inty_m else f"{p.intensity_ma}mA"
                dura_v = dura_m.group(1) if dura_m else f"{p.duration_s}s"
                s_hms  = (wd["start_iso"].split("T")[-1][:8]
                          if "T" in wd["start_iso"] else wd["start_iso"])
                e_hms  = (wd["end_iso"].split("T")[-1][:8]
                          if "T" in wd["end_iso"] else wd["end_iso"])
                sub_fill = PatternFill("solid", fgColor="f0f4ff")

                r_s = _nr()
                ws1.row_dimensions[r_s].height = 15
                c = ws1.cell(r_s, 1, f"  Stimulation {wi + 1}")
                c.font      = Font(bold=True, size=11, color=C1_NAVY)
                c.fill      = sub_fill
                c.alignment = Alignment(horizontal="left", vertical="center")
                ws1.merge_cells(start_row=r_s, start_column=1,
                                end_row=r_s, end_column=4)

                r_s = _nr()
                ws1.row_dimensions[r_s].height = 14
                c = ws1.cell(
                    r_s, 1,
                    f"  Début: t={wd['start_s']:.2f}s ({s_hms})   "
                    f"Fin: t={wd['end_s']:.2f}s ({e_hms})",
                )
                c.font      = Font(size=10, color=C1_LABEL)
                c.fill      = sub_fill
                c.alignment = Alignment(horizontal="left", vertical="center")
                ws1.merge_cells(start_row=r_s, start_column=1,
                                end_row=r_s, end_column=4)

                r_s = _nr()
                ws1.row_dimensions[r_s].height = 14
                c = ws1.cell(
                    r_s, 1,
                    f"  Électrode {elec_v} | Contact {cont_v} | {inty_v} | {dura_v}",
                )
                c.font      = Font(size=10, color=C1_LABEL)
                c.fill      = sub_fill
                c.alignment = Alignment(horizontal="left", vertical="center")
                ws1.merge_cells(start_row=r_s, start_column=1,
                                end_row=r_s, end_column=4)

                if wi < len(stim_detail_windows) - 1:
                    _s1_spacer()

    # ── Section 4: Clinician notes ────────────────────────────────────────────
    _s1_spacer()
    _s1_section_hdr("NOTES CLINICIEN")
    r_note = _nr()
    note_text = clinician_notes or "Aucune note"
    ws1.row_dimensions[r_note].height = max(40, min(120, len(note_text) // 3 + 20))
    c = ws1.cell(r_note, 1, note_text)
    c.font      = Font(
        size=11,
        color=C1_VALUE if clinician_notes else C1_LABEL,
        italic=not bool(clinician_notes),
    )
    c.alignment = Alignment(
        horizontal="left", vertical="top", wrap_text=True, indent=2,
    )
    ws1.merge_cells(start_row=r_note, start_column=1,
                    end_row=r_note, end_column=4)

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 14
    ws1.column_dimensions["C"].width = 16
    ws1.column_dimensions["D"].width = 14

    # ── Sheet 2: Journal des essais ──────────────────────────────────────────
    ws2 = wb.create_sheet("Journal des essais")
    trial_cols = [
        "Essai", "Stimulus", "Heure IMAGE", "Temps IMAGE (s)",
        "Réponse", "Attendue", "Correct", "TR (s)", "Époque", "STIM active", "Notes",
    ]
    ws2.append(trial_cols)
    for cell in ws2[1]:
        cell.font      = _font(bold=True, color="FFFFFF")
        cell.fill      = _fill("1a2744")
        cell.alignment = _center()
        cell.border    = _border()
    ws2.freeze_panes = "A2"

    # Build lookups keyed by essai number
    image_on_map: dict[int, Event] = {}
    response_map: dict[int, Event] = {}
    skip_set: set[int] = set()
    excl_set: set[int] = set()
    for ev in events:
        if ev.event == EventType.IMAGE_ON and ev.essai is not None:
            image_on_map[ev.essai] = ev
        elif ev.event == EventType.RESPONSE and ev.essai is not None:
            response_map[ev.essai] = ev
        elif ev.event == EventType.STIMULUS_SKIP and ev.essai is not None:
            skip_set.add(ev.essai)
        elif ev.event == EventType.STIMULUS_EXCLUDE and ev.essai is not None:
            excl_set.add(ev.essai)

    # Ordered trial list from TRIAL_START events
    trial_numbers: list[int] = []
    _seen: set[int] = set()
    for ev in events:
        if ev.event == EventType.TRIAL_START and ev.essai is not None:
            if ev.essai not in _seen:
                trial_numbers.append(ev.essai)
                _seen.add(ev.essai)

    from openpyxl.styles import Font as _XLFont, Alignment as _XLAlign

    for essai in trial_numbers:
        img_ev  = image_on_map.get(essai)
        resp_ev = response_map.get(essai)

        stim_id    = (resp_ev or img_ev).stimulus if (resp_ev or img_ev) else ""
        stim_label = _format_stimulus_label(stim_id or "", task_code)
        expected   = _expected_response(stim_id or "", task_code)
        img_iso    = img_ev.time_iso if img_ev else ""
        img_ts     = round(img_ev.time_s, 4) if img_ev else ""

        if resp_ev:
            resp_str   = resp_ev.response or ""
            correct_fr = {"Yes": "Oui", "No": "Non"}.get(resp_ev.correct or "", "")
            tr_val     = round(resp_ev.tr_s, 3) if resp_ev.tr_s is not None else ""
            epoch      = resp_ev.stim_epoch or ""
            notes_val  = resp_ev.notes or ""
            if epoch == "per-stim":
                stim_active = "Oui"
            elif epoch == "limite-stim":
                stim_active = "Limite"
            elif epoch:
                stim_active = "Non"
            else:
                stim_active = ""

            if epoch == "per-stim":
                row_hex = "fee2e2"
            elif epoch == "limite-stim":
                row_hex = "fff3e0"
            elif epoch == "post-stim":
                row_hex = "e0f2fe"
            elif resp_ev.correct == "Yes":
                row_hex = "e8f5e9"
            elif resp_ev.correct == "No":
                row_hex = "ffeaea"
            else:
                row_hex = None
        elif essai in skip_set:
            resp_str   = "Passé"
            correct_fr = tr_val = epoch = stim_active = notes_val = ""
            row_hex    = "f5f5f5"
        elif essai in excl_set:
            resp_str   = "Exclu"
            correct_fr = tr_val = epoch = stim_active = notes_val = ""
            row_hex    = "f5f5f5"
        else:
            resp_str   = "Timeout"
            correct_fr = tr_val = epoch = stim_active = notes_val = ""
            row_hex    = "f5f5f5"

        ws2.append([
            essai, stim_label, img_iso, img_ts,
            resp_str, expected, correct_fr, tr_val, epoch, stim_active, notes_val,
        ])
        r = ws2.max_row
        row_fill = _fill(row_hex) if row_hex else None

        for ci, cell in enumerate(ws2[r], start=1):
            cell.border = _border()
            if row_fill:
                cell.fill = row_fill
            # Correct cell — bold colored text
            if ci == 7 and correct_fr:
                cell.font = _font(
                    bold=True,
                    color="16a34a" if correct_fr == "Oui" else "dc2626",
                )
            # TR and Temps IMAGE — right-aligned, monospace
            elif ci in (4, 8):
                cell.font      = _XLFont(name="Courier New", size=10)
                cell.alignment = _XLAlign(horizontal="right", vertical="center")
                continue
            if ci != 7:
                cell.alignment = _left()

    ws2.auto_filter.ref = ws2.dimensions
    for i, w in enumerate([7, 30, 20, 14, 18, 14, 10, 10, 12, 12, 40], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Analyse par stimulation ─────────────────────────────────────
    if n_stim > 0:
        ws4 = wb.create_sheet("Analyse par stimulation")

        # Build stim windows with notes
        stim_windows: list[tuple[float, str, float]] = []
        _open: Optional[tuple[float, str]] = None
        for ev in events:
            if ev.event == EventType.STIM_START:
                _open = (ev.time_s, ev.notes or "")
            elif ev.event == EventType.STIM_END and _open is not None:
                stim_windows.append((_open[0], _open[1], ev.time_s))
                _open = None
        if _open is not None:
            stim_windows.append((_open[0], _open[1], events[-1].time_s))

        # IMAGE_ON lookup
        img_ts_by_essai: dict[int, float] = {
            ev.essai: ev.time_s
            for ev in events
            if ev.event == EventType.IMAGE_ON and ev.essai is not None
        }

        grp_fills  = [_fill("FFFFFF"), _fill("FFE8E8"), _fill("FFF3E0"), _fill("CCEEFF")]
        grp_hdrs   = ["PRÉ-STIM", "PER-STIM", "LIMITE", "POST-STIM"]
        grp_colors = ["000000", "CC0000", "E65100", "0044AA"]
        sub_cols   = ["Essai", "Stimulus", "Réponse", "Correct", "TR (s)"]
        n_cols_per = len(sub_cols)
        total_cols = n_cols_per * 4  # 20

        for wi, (w_start, w_notes, w_end) in enumerate(stim_windows):
            # ── Section header ────────────────────────────────────────────────
            # Parse stim params from notes: "Signal=f12 | Electrode=A Contact=1 ..."
            elec = re.search(r'Electrode=(\S+)', w_notes)
            cont = re.search(r'Contact=(\S+)',   w_notes)
            inty = re.search(r'Intensity=(\S+)', w_notes)
            dura = re.search(r'Duration=(\S+)',  w_notes)
            elec = elec.group(1) if elec else "?"
            cont = cont.group(1) if cont else "?"
            inty = inty.group(1) if inty else "?"
            dura = dura.group(1) if dura else "?"

            hdr_text = (
                f"Stimulation {wi+1} — t={w_start:.2f}s → t={w_end:.2f}s"
                f" | Électrode {elec} | Contact {cont} | {inty} | {dura}"
            )
            r = ws4.max_row + (2 if wi > 0 else 1)
            ws4.cell(r, 1, hdr_text)
            ws4.merge_cells(
                start_row=r, start_column=1, end_row=r, end_column=total_cols
            )
            hdr_cell = ws4.cell(r, 1)
            hdr_cell.font      = _font(bold=True, color="FFFFFF", size=12)
            hdr_cell.fill      = _fill("1a1a2e")
            hdr_cell.alignment = _center()
            hdr_cell.border    = _border()

            # ── Group headers ─────────────────────────────────────────────────
            r += 1
            for gi, (grp, fill, color) in enumerate(
                zip(grp_hdrs, grp_fills, grp_colors)
            ):
                col = gi * n_cols_per + 1
                ws4.cell(r, col, grp)
                ws4.merge_cells(
                    start_row=r, start_column=col,
                    end_row=r, end_column=col + n_cols_per - 1,
                )
                c = ws4.cell(r, col)
                c.font      = _font(bold=True, color=color, size=11)
                c.fill      = fill
                c.alignment = _center()
                c.border    = _border()

            # ── Sub-column headers ────────────────────────────────────────────
            r += 1
            for gi, fill in enumerate(grp_fills):
                for ci, sub in enumerate(sub_cols):
                    col = gi * n_cols_per + ci + 1
                    c = ws4.cell(r, col, sub)
                    c.font      = _font(bold=True, size=10)
                    c.fill      = fill
                    c.alignment = _center()
                    c.border    = _border()

            # ── Classify trials for this window ───────────────────────────────
            prev_end   = stim_windows[wi-1][2] if wi > 0       else 0.0
            next_start = stim_windows[wi+1][0] if wi+1 < len(stim_windows) else float("inf")

            pre_rows: list[Event] = []
            per_rows: list[Event] = []
            lim_rows: list[Event] = []
            post_rows:list[Event] = []

            for resp in resp_evs:
                img_ts = img_ts_by_essai.get(resp.essai) if resp.essai is not None else None
                if img_ts is None:
                    continue
                if img_ts < prev_end or img_ts >= next_start:
                    continue
                epoch = resp.stim_epoch or ""
                if epoch == "per-stim":
                    per_rows.append(resp)
                elif epoch == "limite-stim":
                    lim_rows.append(resp)
                elif epoch == "post-stim":
                    post_rows.append(resp)
                else:
                    pre_rows.append(resp)

            # ── Data rows ─────────────────────────────────────────────────────
            n_data = max(len(pre_rows), len(per_rows), len(lim_rows), len(post_rows))
            for di in range(n_data):
                r += 1
                for gi, (grp_evs, fill, text_color) in enumerate(
                    zip([pre_rows, per_rows, lim_rows, post_rows], grp_fills, grp_colors)
                ):
                    ev = grp_evs[di] if di < len(grp_evs) else None
                    base_col = gi * n_cols_per + 1
                    if ev is not None:
                        correct_fr = {"Yes": "Oui", "No": "Non"}.get(ev.correct or "", "")
                        ok_color   = "007700" if ev.correct == "Yes" else ("CC0000" if ev.correct == "No" else text_color)
                        vals = [
                            ev.essai or "",
                            _format_stimulus_label(ev.stimulus or "", task_code),
                            ev.response or "",
                            correct_fr,
                            round(ev.tr_s, 2) if ev.tr_s is not None else "",
                        ]
                        for ci, val in enumerate(vals):
                            col = base_col + ci
                            c = ws4.cell(r, col, val)
                            c.font      = _font(color=ok_color if ci == 3 else text_color)
                            c.fill      = fill
                            c.alignment = _left()
                            c.border    = _border()
                    else:
                        for ci in range(n_cols_per):
                            c = ws4.cell(r, base_col + ci)
                            c.fill   = fill
                            c.border = _border()

        # Auto-fit columns Sheet 4
        for col_idx in range(1, total_cols + 1):
            max_w = 8
            for row in ws4.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        max_w = max(max_w, len(str(cell.value or "")))
                    except Exception:
                        pass
            ws4.column_dimensions[get_column_letter(col_idx)].width = min(max_w + 2, 32)

        ws4.freeze_panes = "A2"

    # ── Sheet 4: Métadonnées ─────────────────────────────────────────────────
    ws_meta = wb.create_sheet("Métadonnées")
    ws_meta.append(["Clé", "Valeur"])
    for cell in ws_meta[1]:
        cell.font   = _font(bold=True, color="e0e0e0")
        cell.fill   = hdr_fill
        cell.border = _border()
    for k, v in event_log.metadata.items():
        ws_meta.append([k, v])
        for c in ws_meta[ws_meta.max_row]:
            c.border = _border()
    ws_meta.column_dimensions["A"].width = 28
    ws_meta.column_dimensions["B"].width = 60

    # ── Sheet 5: Événements bruts ────────────────────────────────────────────
    ws_raw = wb.create_sheet("Événements bruts")
    raw_cols = ["Time_s", "Time_iso", "Événement", "Essai",
                "Stimulus", "Réponse", "Correct", "TR (s)", "Époque", "Notes"]
    ws_raw.append(raw_cols)
    for cell in ws_raw[1]:
        cell.font      = _font(bold=True, color="e0e0e0")
        cell.fill      = hdr_fill
        cell.alignment = _center()
        cell.border    = _border()

    ev_type_fr = {
        "SESSION_START": "Début session", "TRIAL_START": "Début essai",
        "IMAGE_ON": "Image affichée",     "RESPONSE": "Réponse",
        "STIM_START": "Stim début",       "STIM_END": "Stim fin",
        "TRIAL_END": "Fin essai",         "STIMULUS_SKIP": "Passé",
        "STIMULUS_EXCLUDE": "Exclu",      "STIMULUS_REPLACE": "Remplacé",
        "SESSION_END": "Fin session",     "NOTE": "Note",
    }
    for ev in events:
        stim_label = (
            _format_stimulus_label(ev.stimulus, task_code) if ev.stimulus else ""
        )
        correct_fr = {"Yes": "Oui", "No": "Non"}.get(ev.correct or "", "")
        ws_raw.append([
            round(ev.time_s, 4),
            ev.time_iso,
            ev_type_fr.get(ev.event.value, ev.event.value),
            ev.essai or "",
            stim_label,
            ev.response or "",
            correct_fr,
            round(ev.tr_s, 3) if ev.tr_s is not None else "",
            ev.stim_epoch or "",
            ev.notes or "",
        ])
        row_r = ws_raw.max_row
        if ev.event == EventType.RESPONSE and ev.stim_epoch == "per-stim":
            rf = _fill("FFE8E8")
            for c in ws_raw[row_r]:
                c.fill = rf
        elif ev.event == EventType.RESPONSE and ev.stim_epoch == "limite-stim":
            rf = _fill("FFF3E0")
            for c in ws_raw[row_r]:
                c.fill = rf
        elif ev.event == EventType.RESPONSE and ev.stim_epoch == "post-stim":
            rf = _fill("E8F4FF")
            for c in ws_raw[row_r]:
                c.fill = rf
        for c in ws_raw[row_r]:
            c.border    = _border()
            c.alignment = _left()

    for i, w in enumerate([10, 22, 16, 7, 28, 12, 8, 8, 10, 40], start=1):
        ws_raw.column_dimensions[get_column_letter(i)].width = w
    ws_raw.freeze_panes = "A2"

    wb.save(xl_path)
    return xl_path
